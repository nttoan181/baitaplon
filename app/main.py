
import os, json
from fastapi import FastAPI, UploadFile, File, Request
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import pandas as pd

from app.db import Base, engine, SessionLocal
from app.models import Invoice
from app.ocr_utils import extract_invoice_data

# create tables
Base.metadata.create_all(bind=engine)

app = FastAPI(title='Invoice OCR Dashboard - VN Dark')

app.mount('/static', StaticFiles(directory='app/static'), name='static')
app.mount('/uploads', StaticFiles(directory=os.path.join('data','uploads')), name='uploads')
templates = Jinja2Templates(directory='app/templates')

UPLOAD_FOLDER = os.path.join('data', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.get('/', response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse('landing.html', {'request': request})

@app.get('/dashboard', response_class=HTMLResponse)
def dashboard(request: Request):
    return templates.TemplateResponse('dashboard.html', {'request': request})

@app.post('/api/upload')
async def api_upload(file: UploadFile = File(...)):
    file_path = os.path.join(UPLOAD_FOLDER, file.filename)
    with open(file_path, 'wb') as f:
        f.write(await file.read())

    extracted = extract_invoice_data(file_path)
    total = extracted.get('total_amount')
    try:
        total = float(total) if total is not None else None
    except:
        total = None

    db = SessionLocal()
    inv = Invoice(
        filename=extracted.get('filename'),
        vendor=extracted.get('vendor'),
        tax_id=extracted.get('tax_id'),
        total_amount=total,
        date=extracted.get('date'),
        raw_text=extracted.get('raw_text'),
        parsed_json=json.dumps(extracted.get('parsed_json', {}), ensure_ascii=False)
    )
    db.add(inv)
    db.commit()
    db.refresh(inv)
    db.close()

    return JSONResponse({'status': 'ok', 'invoice': inv.as_dict()})

def query_invoices(db, search=None, vendor=None, tax_id=None, start_date=None, end_date=None, min_total=None, max_total=None, offset=0, limit=50):
    q = db.query(Invoice)
    if search:
        like = f"%{search}%"
        q = q.filter(Invoice.filename.ilike(like) | Invoice.raw_text.ilike(like) | Invoice.vendor.ilike(like) | Invoice.tax_id.ilike(like))
    if vendor:
        q = q.filter(Invoice.vendor.ilike(f"%{vendor}%"))
    if tax_id:
        q = q.filter(Invoice.tax_id.ilike(f"%{tax_id}%"))
    if min_total is not None:
        q = q.filter(Invoice.total_amount >= min_total)
    if max_total is not None:
        q = q.filter(Invoice.total_amount <= max_total)
    if start_date:
        q = q.filter(Invoice.date >= start_date)
    if end_date:
        q = q.filter(Invoice.date <= end_date)
    total = q.count()
    rows = q.order_by(Invoice.created_at.desc()).offset(offset).limit(limit).all()
    return total, rows

@app.get('/api/invoices')
def api_list(q: str = None, vendor: str = None, tax_id: str = None, start_date: str = None, end_date: str = None, min_total: float = None, max_total: float = None, page: int = 1, size: int = 25, invoice_id: int = None):
    db = SessionLocal()
    # exact ID filter shortcut
    if invoice_id is not None:
        qy = db.query(Invoice).filter(Invoice.id == invoice_id)
        total = qy.count()
        rows = qy.offset(0).limit(size).all()
    else:
        offset = (page - 1) * size
        total, rows = query_invoices(db, search=q, vendor=vendor, tax_id=tax_id, start_date=start_date, end_date=end_date, min_total=min_total, max_total=max_total, offset=offset, limit=size)
    items = [r.as_dict() for r in rows]
    db.close()
    return {'total': total, 'page': page, 'size': size, 'items': items}

@app.get('/api/export')
def api_export(q: str = None, vendor: str = None, tax_id: str = None, start_date: str = None, end_date: str = None, min_total: float = None, max_total: float = None, invoice_id: int = None):
    db = SessionLocal()
    if invoice_id is not None:
        rows = db.query(Invoice).filter(Invoice.id == invoice_id).all()
        total = len(rows)
    else:
        total, rows = query_invoices(db, search=q, vendor=vendor, tax_id=tax_id, start_date=start_date, end_date=end_date, min_total=min_total, max_total=max_total, offset=0, limit=1000000)
    items = [r.as_dict() for r in rows]
    db.close()

    headers = ['ID','Tệp','Nhà cung cấp','MST','Tổng tiền','Ngày','Tạo lúc','Văn bản OCR']
    if not items:
        df = pd.DataFrame(columns=headers)
    else:
        df = pd.DataFrame([
            {
                'ID': it['id'],
                'Tệp': it['filename'],
                'Nhà cung cấp': it.get('vendor'),
                'MST': it.get('tax_id'),
                'Tổng tiền': it.get('total_amount'),
                'Ngày': it.get('date'),
                'Tạo lúc': it.get('created_at'),
                'Văn bản OCR': it.get('raw_text'),
            } for it in items
        ])

    out_path = os.path.join('data', 'invoices_export.xlsx')
    # save xlsx and then style header/columns for readability
    # normalize Unicode to NFC for all string cells
    import unicodedata
    df = df.applymap(lambda v: unicodedata.normalize('NFC', v) if isinstance(v, str) else v)
    df.to_excel(out_path, index=False)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment
        wb = load_workbook(out_path)
        font = Font(name='Arial')
        for ws in wb.worksheets:
            # freeze header
            ws.freeze_panes = 'A2'
            # bold header
            for cell in ws[1]:
                cell.font = Font(name='Arial', bold=True)
            # wrap OCR column and set widths
            ocr_col = None
            for idx, cell in enumerate(ws[1], start=1):
                if str(cell.value).strip().lower() == 'văn bản ocr':
                    ocr_col = idx
                # auto-ish width
                header_len = len(str(cell.value)) if cell.value else 10
                ws.column_dimensions[cell.column_letter].width = max(12, min(32, header_len + 6))
            # widen OCR column specifically
            if ocr_col:
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(ocr_col)].width = 60
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font
                    if ocr_col and cell.column == ocr_col:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        wb.save(out_path)
    except Exception:
        # fallback: keep default
        pass
    return FileResponse(out_path, filename='invoices_export.xlsx', media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get('/api/invoice/{invoice_id}')
def api_get(invoice_id: int):
    db = SessionLocal()
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    db.close()
    if not inv:
        return JSONResponse({'error': 'not found'}, status_code=404)
    return inv.as_dict()

@app.delete('/api/invoice/{invoice_id}')
def api_delete(invoice_id: int):
    db = SessionLocal()
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not inv:
        db.close()
        return JSONResponse({'error': 'not found'}, status_code=404)
    
    # Delete the uploaded file if it exists
    try:
        if inv.filename:
            file_path = os.path.join(UPLOAD_FOLDER, inv.filename)
            if os.path.exists(file_path):
                os.remove(file_path)
    except Exception as e:
        print(f"Warning: Could not delete file {inv.filename}: {e}")
    
    db.delete(inv)
    db.commit()
    db.close()
    return JSONResponse({'status': 'ok', 'message': 'Invoice deleted successfully'})

@app.get('/invoice/{invoice_id}', response_class=HTMLResponse)
def view_invoice(request: Request, invoice_id: int):
    db = SessionLocal()
    inv = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    db.close()
    if not inv:
        return HTMLResponse('<h1>Không tìm thấy</h1>', status_code=404)
    data = inv.as_dict()
    fname = data.get('filename') or ''
    lower = fname.lower()
    is_image = any(lower.endswith(ext) for ext in ['.png','.jpg','.jpeg','.webp','.gif'])
    image_url = f"/uploads/{fname}" if is_image else None
    return templates.TemplateResponse('invoice_detail.html', {'request': request, 'invoice': data, 'image_url': image_url})
